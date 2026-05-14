"""Build the final Option C output Excel.

Consumes:
    - input/indicators.xlsx (the codebook)
    - input/scores.xlsx (scores + analysis from Claude Code)
    - input/states.yaml (state metadata)

Produces:
    - output/State_Constitutional_Analysis.xlsx

The Excel has the Option C hybrid tidy + presentation structure:
    1. Read Me First
    2. Executive Summary
    3. FF — Comparative
    4. CB — Comparative
    5. DP — Comparative
    6. Tidy Data
    7. Analysis (narrative from Claude Code)
    8. Sources
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


log = logging.getLogger(__name__)

# ============== Style constants ==============
FONT_NAME = "Arial"

COLOR_HEADER_BG = "1F3864"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUBHEADER_BG = "BDD7EE"
COLOR_NOTE_BG = "F2F2F2"
COLOR_COVER_TITLE = "1F3864"

DIM_COLORS = {
    "Fundamental Freedoms": "DEEBF7",
    "Checks and Balances": "FFF2CC",
    "Democratic Protections": "E2EFDA",
}

SCORE_COLORS = {
    0: "E06666",
    1: "F6B26B",
    2: "FFE599",
    3: "93C47D",
}

DIMENSIONS = ["Fundamental Freedoms", "Checks and Balances", "Democratic Protections"]

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# ============== Style helpers ==============

def header_style(cell):
    cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FG, size=11)
    cell.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def body_style(cell, wrap=True, align="left"):
    cell.font = Font(name=FONT_NAME, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = THIN_BORDER


def score_style(cell, score):
    cell.font = Font(name=FONT_NAME, bold=True, size=12)
    if isinstance(score, (int, float)) and not isinstance(score, bool):
        try:
            color = SCORE_COLORS[int(score)]
            cell.fill = PatternFill("solid", start_color=color)
        except (KeyError, ValueError):
            pass
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def set_col_widths(ws, widths: dict[str, int]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ============== Data loading ==============

@dataclass
class Indicator:
    id: str
    dimension: str
    indicator: str
    question: str
    scoring_guidance: str


@dataclass
class Score:
    indicator_id: str
    state_code: str
    score: Any  # int, None, or sentinel
    constitutional_reference: str
    note: str


@dataclass
class AnalysisSection:
    section_id: str
    content: str


@dataclass
class State:
    code: str
    name: str


def load_indicators(path: Path) -> list[Indicator]:
    df = pd.read_excel(path, sheet_name="indicators")
    required = {"id", "dimension", "indicator", "question", "scoring_guidance"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"indicators.xlsx missing columns: {missing}")
    out = []
    for _, row in df.iterrows():
        out.append(
            Indicator(
                id=str(row["id"]).strip(),
                dimension=str(row["dimension"]).strip(),
                indicator=str(row["indicator"]).strip(),
                question=str(row["question"]).strip(),
                scoring_guidance=str(row["scoring_guidance"]).strip(),
            )
        )
    return out


def load_scores(path: Path) -> tuple[list[Score], list[AnalysisSection]]:
    """Load scores and analysis from the Claude Code-generated input/scores.xlsx.

    If the file does not exist, returns empty lists (the builder will still
    produce a structurally valid workbook with blank scores).
    """
    if not path.exists():
        log.warning("scores.xlsx not found at %s; output will have blank scores", path)
        return [], []

    scores_df = pd.read_excel(path, sheet_name="scores")
    required = {"indicator_id", "state_code", "score", "constitutional_reference", "note"}
    missing = required - set(scores_df.columns)
    if missing:
        raise ValueError(f"scores.xlsx sheet 'scores' missing columns: {missing}")

    scores = []
    for _, row in scores_df.iterrows():
        # Handle possible NaN/None values gracefully
        sc = row["score"]
        if pd.isna(sc):
            sc_val = None
        else:
            try:
                sc_val = int(sc)
            except (ValueError, TypeError):
                sc_val = None
        scores.append(
            Score(
                indicator_id=str(row["indicator_id"]).strip(),
                state_code=str(row["state_code"]).strip(),
                score=sc_val,
                constitutional_reference=str(row["constitutional_reference"]).strip() if pd.notna(row["constitutional_reference"]) else "",
                note=str(row["note"]).strip() if pd.notna(row["note"]) else "",
            )
        )

    try:
        analysis_df = pd.read_excel(path, sheet_name="analysis")
        analysis = []
        for _, row in analysis_df.iterrows():
            if pd.isna(row.get("section_id")) or pd.isna(row.get("content")):
                continue
            analysis.append(
                AnalysisSection(
                    section_id=str(row["section_id"]).strip(),
                    content=str(row["content"]).strip(),
                )
            )
    except (KeyError, ValueError):
        log.warning("scores.xlsx has no 'analysis' sheet; output will lack narrative")
        analysis = []

    return scores, analysis


def load_states(path: Path) -> list[State]:
    with open(path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return [State(code=s["code"], name=s["name"]) for s in data["states"]]


# ============== Score lookup helper ==============

class ScoreLookup:
    def __init__(self, scores: list[Score]):
        self._map = {(s.indicator_id, s.state_code): s for s in scores}

    def get(self, indicator_id: str, state_code: str) -> Score:
        key = (indicator_id, state_code)
        if key in self._map:
            return self._map[key]
        # Return a blank score for missing combinations
        return Score(
            indicator_id=indicator_id,
            state_code=state_code,
            score=None,
            constitutional_reference="",
            note="",
        )


# ============== Sheet builders ==============

GLOSSARY_TEXT = [
    ("Purpose", "This workbook contains a de jure comparative analysis of U.S. state constitutions to support pilot state selection for a study on Fundamental Freedoms, Checks and Balances, and Democratic Protections."),
    ("Scope", "This is a TEXTUAL analysis (de jure). It does not capture: (a) judicial interpretation, (b) statutory implementation, (c) enforcement practice, or (d) political dynamics. A clause in the text may be unenforceable or interpreted differently from its plain reading."),
    ("Dimensions", "(1) Fundamental Freedoms — individual rights and protections; (2) Checks and Balances — institutional accountability among branches; (3) Democratic Protections — popular sovereignty mechanisms and electoral rights."),
    ("Scoring", "Ordinal 0–3 scale where HIGHER = stronger democratic protection. Direction defined per indicator in the 'Scoring guidance' field."),
    ("Score meaning", "0 = Absent, regressive, or strongly weakening. 1 = Weak or partially compromised. 2 = Moderate or standard. 3 = Strong or expansive."),
    ("Aggregation", "Aggregate scores are simple sums with equal weighting — a deliberate simplification for transparency. To use weighted scoring, recompute from the 'Tidy Data' sheet."),
    ("Reproducibility", "Scores produced by Claude Code reading the constitutional text. Once generated, the canonical scores are in input/scores.xlsx and can be manually edited. The output Excel is rebuilt deterministically by code/02_build_output.py."),
    ("Use case", "Suitable for: comparing states' formal constitutional profiles, identifying contrasts useful for case selection, generating hypotheses. NOT suitable for: standalone policy advocacy, normative ranking, or predicting practice."),
]


def build_cover_sheet(wb: Workbook, subtitle: str):
    ws = wb.active
    ws.title = "Read Me First"

    ws["B2"] = "State Constitutional Analysis"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=22, color=COLOR_COVER_TITLE)

    ws["B4"] = subtitle
    ws["B4"].font = Font(name=FONT_NAME, italic=True, size=12, color="555555")

    ws["B6"] = "Generated by code/02_build_output.py"
    ws["B6"].font = Font(name=FONT_NAME, size=10, color="888888")

    row = 9
    for title, text in GLOSSARY_TEXT:
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name=FONT_NAME, bold=True, size=11, color=COLOR_HEADER_BG)
        c.alignment = Alignment(vertical="top")

        c = ws.cell(row=row, column=3, value=text)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(40, len(text) // 10)
        row += 1

    set_col_widths(ws, {"A": 2, "B": 26, "C": 95, "D": 2})


def build_executive_summary(
    wb: Workbook,
    indicators: list[Indicator],
    states: list[State],
    sheet_locations: dict[str, tuple[str, int]],
):
    """Build executive summary referencing aggregate cells from dimension sheets.

    sheet_locations: dict mapping dimension name -> (sheet_name, aggregate_row).
    """
    ws = wb.create_sheet("Executive Summary")

    ws["B2"] = "Executive Summary — Aggregate Scores by Dimension"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_COVER_TITLE)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + len(states) + 1)

    intro = (
        "Each cell shows the sum of indicator scores within that dimension for that state. "
        "Higher = stronger democratic profile (per the scoring direction defined per indicator). "
        "Aggregates use equal weighting; use the Tidy Data sheet to apply your own weights."
    )
    ws["B4"] = intro
    ws["B4"].font = Font(name=FONT_NAME, italic=True, size=10, color="555555")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=2 + len(states) + 1)
    ws.row_dimensions[4].height = 40

    # Header row
    headers = ["Dimension"] + [s.code for s in states] + ["Max"]
    for i, h in enumerate(headers):
        c = ws.cell(row=6, column=2 + i, value=h)
        header_style(c)
    ws.row_dimensions[6].height = 24

    # Count max per dimension
    dim_count = {}
    for ind in indicators:
        dim_count[ind.dimension] = dim_count.get(ind.dimension, 0) + 1

    # Rows
    state_col_letters = {
        s.code: get_column_letter(3 + i)  # state cols start at column 3 in dimension sheets
        for i, s in enumerate(states)
    }

    row = 7
    for dim in DIMENSIONS:
        c = ws.cell(row=row, column=2, value=dim)
        c.font = Font(name=FONT_NAME, bold=True, size=11)
        c.fill = PatternFill("solid", start_color=DIM_COLORS[dim])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = THIN_BORDER

        sheet_name, agg_row = sheet_locations[dim]
        for j, state in enumerate(states):
            col_letter = state_col_letters[state.code]
            formula = f"='{sheet_name}'!{col_letter}{agg_row}"
            c = ws.cell(row=row, column=3 + j, value=formula)
            c.font = Font(name=FONT_NAME, bold=True, size=12)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER

        max_val = dim_count.get(dim, 0) * 3
        c = ws.cell(row=row, column=3 + len(states), value=max_val)
        c.font = Font(name=FONT_NAME, italic=True, size=10, color="555555")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

        ws.row_dimensions[row].height = 26
        row += 1

    # Total row
    c = ws.cell(row=row, column=2, value="TOTAL (de jure profile)")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = THIN_BORDER

    for j in range(len(states)):
        col_letter = get_column_letter(3 + j)
        formula = f"=SUM({col_letter}7:{col_letter}{row - 1})"
        c = ws.cell(row=row, column=3 + j, value=formula)
        c.font = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    total_max = sum(dim_count.values()) * 3
    c = ws.cell(row=row, column=3 + len(states), value=total_max)
    c.font = Font(name=FONT_NAME, bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 28

    # Color scale
    score_range = f"C7:{get_column_letter(2 + len(states))}{row - 1}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="E06666",
            mid_type="num", mid_value=10, mid_color="FFE599",
            end_type="num", end_value=21, end_color="93C47D",
        ),
    )

    # Column widths
    widths = {"A": 2, "B": 36}
    for i, s in enumerate(states):
        widths[get_column_letter(3 + i)] = 11
    widths[get_column_letter(3 + len(states))] = 10
    widths[get_column_letter(4 + len(states))] = 2
    set_col_widths(ws, widths)

    ws.freeze_panes = "C7"


def build_comparative_sheet(
    wb: Workbook,
    dimension: str,
    short_code: str,
    indicators: list[Indicator],
    states: list[State],
    lookup: ScoreLookup,
) -> tuple[str, int]:
    """Build one comparative sheet. Returns (sheet_name, aggregate_row)."""
    sheet_name = f"{short_code} — Comparative"
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = f"{dimension} — De jure comparative"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_COVER_TITLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(states) + 2)

    # Column structure:
    # A=ID, B=Indicator, then states (one column each), then Question, Scoring guidance
    n_states = len(states)
    headers = ["ID", "Indicator"] + [s.code for s in states] + ["Question", "Scoring guidance"]
    for i, h in enumerate(headers):
        c = ws.cell(row=3, column=1 + i, value=h)
        header_style(c)
    ws.row_dimensions[3].height = 28

    dim_indicators = [ind for ind in indicators if ind.dimension == dimension]

    row = 4
    for ind in dim_indicators:
        # Score row
        c = ws.cell(row=row, column=1, value=ind.id)
        body_style(c, align="center")
        c.font = Font(name=FONT_NAME, bold=True, size=10)

        c = ws.cell(row=row, column=2, value=ind.indicator)
        body_style(c)
        c.font = Font(name=FONT_NAME, bold=True, size=10)

        for j, state in enumerate(states):
            sc = lookup.get(ind.id, state.code)
            val = sc.score if sc.score is not None else ""
            c = ws.cell(row=row, column=3 + j, value=val)
            score_style(c, sc.score)

        c = ws.cell(row=row, column=3 + n_states, value=ind.question)
        body_style(c)
        c = ws.cell(row=row, column=4 + n_states, value=ind.scoring_guidance)
        body_style(c)

        ws.row_dimensions[row].height = 55

        # Reference row
        ref_row = row + 1
        ws.cell(row=ref_row, column=1).border = THIN_BORDER
        c = ws.cell(row=ref_row, column=2, value="Constitutional reference:")
        c.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = THIN_BORDER

        for j, state in enumerate(states):
            sc = lookup.get(ind.id, state.code)
            c = ws.cell(row=ref_row, column=3 + j, value=sc.constitutional_reference)
            c.font = Font(name=FONT_NAME, italic=True, size=9, color="333333")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", start_color=COLOR_NOTE_BG)
            c.border = THIN_BORDER

        ws.cell(row=ref_row, column=3 + n_states).border = THIN_BORDER
        ws.cell(row=ref_row, column=4 + n_states).border = THIN_BORDER
        ws.row_dimensions[ref_row].height = 28

        # Note row
        note_row = row + 2
        ws.cell(row=note_row, column=1).border = THIN_BORDER
        c = ws.cell(row=note_row, column=2, value="Notes:")
        c.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
        c.alignment = Alignment(horizontal="right", vertical="top")
        c.border = THIN_BORDER

        for j, state in enumerate(states):
            sc = lookup.get(ind.id, state.code)
            c = ws.cell(row=note_row, column=3 + j, value=sc.note)
            c.font = Font(name=FONT_NAME, size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.fill = PatternFill("solid", start_color=COLOR_NOTE_BG)
            c.border = THIN_BORDER

        ws.cell(row=note_row, column=3 + n_states).border = THIN_BORDER
        ws.cell(row=note_row, column=4 + n_states).border = THIN_BORDER
        ws.row_dimensions[note_row].height = 90

        row += 3

    # Aggregate row
    agg_row = row + 1
    ws.cell(row=agg_row, column=1).border = THIN_BORDER
    c = ws.cell(row=agg_row, column=2, value=f"Aggregate {short_code} score (sum)")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = THIN_BORDER

    score_rows = [4 + 3 * k for k in range(len(dim_indicators))]
    for j in range(n_states):
        col_letter = get_column_letter(3 + j)
        cells = ",".join(f"{col_letter}{r}" for r in score_rows)
        formula = f"=SUM({cells})"
        c = ws.cell(row=agg_row, column=3 + j, value=formula)
        c.font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    max_score = len(dim_indicators) * 3
    c = ws.cell(row=agg_row, column=3 + n_states, value=f"Max: {max_score}")
    c.font = Font(name=FONT_NAME, italic=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    ws.merge_cells(start_row=agg_row, start_column=3 + n_states, end_row=agg_row, end_column=4 + n_states)
    ws.row_dimensions[agg_row].height = 30

    # Column widths
    widths = {"A": 7, "B": 38}
    for i, s in enumerate(states):
        widths[get_column_letter(3 + i)] = 9
    widths[get_column_letter(3 + n_states)] = 38
    widths[get_column_letter(4 + n_states)] = 42
    set_col_widths(ws, widths)

    ws.freeze_panes = f"{get_column_letter(3)}4"

    return sheet_name, agg_row


def build_tidy_data_sheet(
    wb: Workbook,
    indicators: list[Indicator],
    states: list[State],
    lookup: ScoreLookup,
):
    ws = wb.create_sheet("Tidy Data")

    ws["A1"] = "Tidy Data — Long format (one row per indicator × state)"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_COVER_TITLE)
    ws.merge_cells("A1:I1")

    ws["A2"] = "Structured for analytical use: pivot tables, R/Python imports, custom weighted aggregations."
    ws["A2"].font = Font(name=FONT_NAME, italic=True, size=10, color="555555")
    ws.merge_cells("A2:I2")

    headers = ["indicator_id", "dimension", "indicator", "state_code", "state_name", "score", "constitutional_reference", "note", "scoring_guidance"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=1 + i, value=h)
        header_style(c)
    ws.row_dimensions[4].height = 26

    state_name_map = {s.code: s.name for s in states}

    row = 5
    for ind in indicators:
        for state in states:
            sc = lookup.get(ind.id, state.code)

            ws.cell(row=row, column=1, value=ind.id)
            body_style(ws.cell(row=row, column=1), align="center")
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=10, bold=True)

            c = ws.cell(row=row, column=2, value=ind.dimension)
            body_style(c)
            c.fill = PatternFill("solid", start_color=DIM_COLORS[ind.dimension])

            body_style(ws.cell(row=row, column=3, value=ind.indicator))
            body_style(ws.cell(row=row, column=4, value=state.code), align="center")
            ws.cell(row=row, column=4).font = Font(name=FONT_NAME, size=10, bold=True)
            body_style(ws.cell(row=row, column=5, value=state.name))

            c = ws.cell(row=row, column=6, value=sc.score if sc.score is not None else "")
            score_style(c, sc.score)

            c = ws.cell(row=row, column=7, value=sc.constitutional_reference)
            body_style(c)
            c.font = Font(name=FONT_NAME, italic=True, size=9)

            c = ws.cell(row=row, column=8, value=sc.note)
            body_style(c)
            c.font = Font(name=FONT_NAME, size=9)

            c = ws.cell(row=row, column=9, value=ind.scoring_guidance)
            body_style(c)
            c.font = Font(name=FONT_NAME, size=9, italic=True, color="555555")

            ws.row_dimensions[row].height = 45
            row += 1

    set_col_widths(ws, {
        "A": 10, "B": 22, "C": 36, "D": 8, "E": 18, "F": 8, "G": 24, "H": 60, "I": 45,
    })
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{row - 1}"


def build_analysis_sheet(wb: Workbook, analysis: list[AnalysisSection]):
    """Build the narrative analysis sheet from the analysis sections."""
    ws = wb.create_sheet("Analysis")
    ws["B2"] = "Narrative Analysis"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_COVER_TITLE)

    intro = (
        "Generated by Claude Code during evaluation. Each section was written based on "
        "reading the constitutional texts and the scoring results. Treat as input to "
        "discussion, not as a final ranking or recommendation."
    )
    ws["B3"] = intro
    ws["B3"].font = Font(name=FONT_NAME, italic=True, size=10, color="555555")
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 40

    if not analysis:
        ws["B5"] = "No analysis content provided. Did Claude Code generate the 'analysis' sheet in input/scores.xlsx?"
        ws["B5"].font = Font(name=FONT_NAME, italic=True, color="888888")
        return

    # Define preferred section ordering and human-readable labels
    section_titles = {
        "executive_summary": "Executive Summary",
        "key_pattern": "Key Comparative Pattern",
        "profile_SC": "Profile: South Carolina",
        "profile_TN": "Profile: Tennessee",
        "profile_TX": "Profile: Texas",
        "profile_UT": "Profile: Utah",
        "profile_MN": "Profile: Minnesota",
        "profile_OH": "Profile: Ohio",
        "selection_guidance": "Pilot Selection Guidance",
    }
    preferred_order = list(section_titles.keys())

    sections_by_id = {s.section_id: s for s in analysis}
    ordered = []
    for sid in preferred_order:
        if sid in sections_by_id:
            ordered.append(sections_by_id[sid])
    # Append any sections not in the preferred ordering, in order they appeared
    for s in analysis:
        if s.section_id not in preferred_order:
            ordered.append(s)

    row = 5
    for s in ordered:
        title = section_titles.get(s.section_id, s.section_id)
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(name=FONT_NAME, bold=True, size=12, color=COLOR_HEADER_BG)
        c.alignment = Alignment(vertical="top")

        c = ws.cell(row=row, column=3, value=s.content)
        c.font = Font(name=FONT_NAME, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(60, min(300, len(s.content) // 4))
        row += 1

    set_col_widths(ws, {"A": 2, "B": 28, "C": 95, "D": 2})


SOURCES = [
    ("50 Constitutions", "State Democracy Research Initiative, UW Law School", "https://50constitutions.org/"),
    ("South Carolina Constitution", "SC Legislature", "https://www.scstatehouse.gov/scconstitution/scconst.php"),
    ("Tennessee Constitution", "TN Secretary of State", "https://publications.tnsosfiles.com/pub/2023%20TN%20Constitution.pdf"),
    ("Texas Constitution", "Texas Legislature Online", "https://statutes.capitol.texas.gov/?link=CN"),
    ("Utah Constitution", "Utah State Legislature", "https://le.utah.gov/xcode/constitution.html"),
    ("Minnesota Constitution", "MN Office of the Revisor of Statutes", "https://www.revisor.mn.gov/constitution/"),
    ("Ohio Constitution", "Ohio Laws and Administrative Rules", "https://codes.ohio.gov/ohio-constitution"),
    ("Ballotpedia state constitution pages", "Ballotpedia", "https://ballotpedia.org/"),
    ("Comparative Constitutions Project (CCP) Codebook v4.0", "Elkins, Z. & Ginsburg, T. (2022)", "https://comparativeconstitutionsproject.org/"),
    ("Expanding the Vote: State Felony Disenfranchisement Reform", "The Sentencing Project", "https://www.sentencingproject.org/"),
    ("Brennan Center for Justice — judicial selection research", "Brennan Center", "https://www.brennancenter.org/"),
    ("State Court Report", "Brennan Center / NYU Law", "https://statecourtreport.org/"),
]


def build_sources_sheet(wb: Workbook):
    ws = wb.create_sheet("Sources")
    ws["B2"] = "Sources and References"
    ws["B2"].font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_COVER_TITLE)

    headers = ["#", "Source", "Publisher / Author", "URL"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=2 + i, value=h)
        header_style(c)
    ws.row_dimensions[4].height = 24

    for i, (source, publisher, url) in enumerate(SOURCES, start=1):
        ws.cell(row=4 + i, column=2, value=i)
        body_style(ws.cell(row=4 + i, column=2), align="center")
        body_style(ws.cell(row=4 + i, column=3, value=source))
        body_style(ws.cell(row=4 + i, column=4, value=publisher))
        c = ws.cell(row=4 + i, column=5, value=url)
        c.font = Font(name=FONT_NAME, size=9, color="0563C1", underline="single")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = THIN_BORDER
        c.hyperlink = url
        ws.row_dimensions[4 + i].height = 30

    set_col_widths(ws, {"A": 2, "B": 5, "C": 50, "D": 40, "E": 70})


# ============== Top-level builder ==============

def build_workbook(
    indicators: list[Indicator],
    states: list[State],
    scores: list[Score],
    analysis: list[AnalysisSection],
    output_path: Path,
):
    """Build the full Option C workbook and save."""
    wb = Workbook()

    n_states_str = f"{len(states)} states: {', '.join(s.code for s in states)}"
    build_cover_sheet(
        wb,
        subtitle=f"De jure comparative analysis — {n_states_str} — Hybrid Tidy + Presentation",
    )

    lookup = ScoreLookup(scores)

    # We need executive summary first but it references the comparative sheets.
    # Strategy: create executive summary placeholder, then comparatives, then patch
    # exec summary with formulas (since openpyxl writes formulas as strings, we can
    # build exec summary FIRST after computing aggregate row positions).
    #
    # Compute aggregate row positions a priori for each dimension.
    sheet_locations: dict[str, tuple[str, int]] = {}
    short_codes = {"Fundamental Freedoms": "FF", "Checks and Balances": "CB", "Democratic Protections": "DP"}
    for dim in DIMENSIONS:
        n = sum(1 for ind in indicators if ind.dimension == dim)
        # Each indicator takes 3 rows starting at row 4. Aggregate row = 4 + 3*n + 1
        agg_row = 4 + 3 * n + 1
        sheet_locations[dim] = (f"{short_codes[dim]} — Comparative", agg_row)

    # Now build in the actual desired order
    build_executive_summary(wb, indicators, states, sheet_locations)
    for dim in DIMENSIONS:
        build_comparative_sheet(wb, dim, short_codes[dim], indicators, states, lookup)
    build_tidy_data_sheet(wb, indicators, states, lookup)
    build_analysis_sheet(wb, analysis)
    build_sources_sheet(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Wrote output workbook: %s", output_path)
    return output_path
