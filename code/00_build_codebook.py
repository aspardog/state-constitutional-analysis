"""Generate the editable codebook (input/indicators.xlsx).

This script is used once during project setup. After generation,
edit indicators.xlsx by hand if needed.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---- The codebook content ----
INDICATORS = [
    # FUNDAMENTAL FREEDOMS
    ("FF01", "Fundamental Freedoms", "Rights placement and structural prominence",
     "Where in the constitution is the Declaration of Rights located, and how prominently?",
     "0=Absent or buried late in the document; 1=Present but not in Article I; 2=In Article I, basic enumeration; 3=In Article I, comprehensive, placed BEFORE government structure articles"),

    ("FF02", "Fundamental Freedoms", "Equal protection / non-discrimination clause",
     "Does the constitution have an explicit equal protection or non-discrimination clause?",
     "0=No clause; 1=General reference to equality; 2=Explicit equal protection clause; 3=Explicit clause with enumerated protected categories (gender, race, etc.)"),

    ("FF03", "Fundamental Freedoms", "Privacy right (explicit or interpretively recognized)",
     "Does the constitution recognize a right to privacy, explicitly or via clear textual basis?",
     "0=No textual basis for privacy; 1=Implicit only via search and seizure / due process; 2=Implicit but with broader textual hooks; 3=Explicit textual right to privacy"),

    ("FF04", "Fundamental Freedoms", "Religious freedom and church-state separation",
     "How strong is the religious freedom guarantee and church-state separation in the text?",
     "0=State religion specified or strong establishment; 1=Freedom guaranteed but with qualifiers; 2=Freedom plus secular framing; 3=Freedom plus explicit separation clause, no exclusionary or qualifying language"),

    ("FF05", "Fundamental Freedoms", "Religious test for public office (textual exclusion)",
     "Does the constitutional text contain any religious test or exclusion for public office?",
     "0=Active exclusion of non-believers or specific groups in text; 1=Qualifier present (e.g., must believe in Supreme Being); 2=Test prohibited but with historical caveats remaining; 3=Clean prohibition, no exclusionary text"),

    ("FF06", "Fundamental Freedoms", "Right to bail and pre-trial release",
     "How robust is the right to bail / pre-trial release in the constitutional text?",
     "0=No right or extensive exceptions; 1=Right with documented progressive textual erosion; 2=Right with standard exceptions (capital cases only); 3=Strong right with narrow exceptions only"),

    ("FF07", "Fundamental Freedoms", "Constitutional treatment of abortion / reproductive rights",
     "Does the constitution explicitly address (protect, exclude, or remain silent on) reproductive rights?",
     "0=Explicitly excluded as a right; 1=Silent and narrowly interpreted by courts; 2=Silent, not foreclosed by text; 3=Explicitly protected"),

    # CHECKS AND BALANCES
    ("CB01", "Checks and Balances", "Plural executive (independently elected statewide officials)",
     "How many statewide executive officials are independently elected (not appointed by the Governor)?",
     "0=Single executive (Governor only); 1=2-3 officials; 2=4-5 officials; 3=6+ officials (highly fragmented executive)"),

    ("CB02", "Checks and Balances", "Attorney General independence from Governor",
     "How is the state Attorney General selected, and how independent from the Governor?",
     "0=Appointed by Governor (no independence); 1=Appointed by legislature; 2=Appointed by judiciary (independent from Governor but captured elsewhere); 3=Independently elected by voters"),

    ("CB03", "Checks and Balances", "Judicial selection method (high court)",
     "How are state supreme court justices selected?",
     "0=Legislative appointment (captured by legislature); 1=Partisan elections (politicized); 2=Gubernatorial appointment alone; 3=Merit selection commission + retention (insulated from raw political pressure)"),

    ("CB04", "Checks and Balances", "Judicial independence from legislative branch",
     "Does the legislative branch dominate judicial selection or judicial discipline?",
     "0=Legislature controls both selection and discipline; 1=Legislature confirms appointees with strong role; 2=Limited legislative role (e.g., Senate confirmation only); 3=No legislative role in selection or discipline"),

    ("CB05", "Checks and Balances", "Lieutenant Governor power and accountability",
     "How is the Lt Governor selected and how much cross-branch power does the office concentrate?",
     "0=Lt Gov is Senate Speaker (fusion of powers) OR independently elected with major legislative/executive powers concentrated; 1=Independently elected but modest role; 2=Joint ticket with Governor, ceremonial role; 3=Joint ticket, modest succession-only role with no cross-branch concentration"),

    ("CB06", "Checks and Balances", "Veto override threshold",
     "What legislative threshold is required to override a gubernatorial veto?",
     "0=Simple majority (weakest veto); 1=3/5 majority; 2=2/3 majority (standard); 3=3/4 or higher"),

    ("CB07", "Checks and Balances", "Constitutional judicial conduct/discipline body",
     "Is there a constitutionally established body to discipline judges?",
     "0=No body or only legislative impeachment available; 1=Statutory body only (not constitutional); 2=Constitutional body with limited powers; 3=Constitutional body with comprehensive discipline powers"),

    ("CB08", "Checks and Balances", "Single-subject rule for legislation",
     "Does the constitution require each law/amendment to address only a single subject?",
     "0=No rule; 1=Weak or narrowly applied; 2=Applies to legislation; 3=Applies to both legislation AND constitutional amendments"),

    # DEMOCRATIC PROTECTIONS
    ("DP01", "Democratic Protections", "Citizen initiative for statutes",
     "Can citizens propose statutes through signature petition (statutory initiative)?",
     "0=No initiative process; 1=Yes, but with very high signature threshold or major restrictions; 2=Yes, moderate threshold; 3=Yes, accessible threshold"),

    ("DP02", "Democratic Protections", "Citizen initiative for constitutional amendments",
     "Can citizens propose constitutional amendments directly via signature petition?",
     "0=No; 1=Yes with very high threshold; 2=Yes, moderate; 3=Yes, accessible"),

    ("DP03", "Democratic Protections", "Veto referendum (citizen check on legislation)",
     "Can citizens petition to suspend/repeal legislative acts via referendum?",
     "0=No; 1=Yes with very high threshold; 2=Yes, moderate; 3=Yes, accessible"),

    ("DP04", "Democratic Protections", "Recall of statewide elected officials",
     "Can voters recall statewide elected officials via petition?",
     "0=No recall mechanism; 1=Local officials only; 2=Limited statewide (specific offices only); 3=Full statewide recall available"),

    ("DP05", "Democratic Protections", "Constitutional amendment legislative threshold",
     "How strict is the legislative threshold to refer constitutional amendments to voters?",
     "Note: higher rigidity entrenches existing text (both rights and restrictions). 0=Simple majority; 1=3/5; 2=2/3 in single session; 3=2/3 across two consecutive sessions or higher bar"),

    ("DP06", "Democratic Protections", "Constitutional amendment popular ratification threshold",
     "What popular threshold is required to ratify a constitutional amendment?",
     "0=No popular vote required; 1=Simple majority of votes on the amendment; 2=Majority of votes cast in the election; 3=Majority of votes for top-of-ticket office (very strict denominator)"),

    ("DP07", "Democratic Protections", "Felony disenfranchisement strictness",
     "How restrictively does the state disenfranchise people with felony convictions?",
     "0=Lifetime ban for any felony; 1=Until completion of full sentence including parole and probation (and beyond); 2=Until completion of incarceration plus parole/probation; 3=Only during incarceration; restoration automatic upon release"),
]


def build_codebook(output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "indicators"

    # Header row
    headers = ["id", "dimension", "indicator", "question", "scoring_guidance"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(INDICATORS, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 60

    # Column widths
    widths = {"A": 8, "B": 24, "C": 38, "D": 50, "E": 70}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(INDICATORS)+1}"

    # Add a second sheet with documentation
    doc = wb.create_sheet("documentation")
    doc_lines = [
        ("Purpose", "This file is the editable codebook for the State Constitutional Analysis pipeline."),
        ("Editing", "Add or remove rows in the 'indicators' sheet. Keep 'id' unique and stable (do not renumber)."),
        ("Dimensions", "Use exactly one of: 'Fundamental Freedoms', 'Checks and Balances', 'Democratic Protections'."),
        ("Scoring", "All indicators use a 0–3 ordinal scale. Higher = stronger democratic protection. Define direction in scoring_guidance."),
        ("After editing", "Run Claude Code's /analyze command to regenerate scores for new or changed indicators. Then re-run 'python code/02_build_output.py'."),
        ("Original 22 indicators", "Designed for SC/TN/TX/UT/MN/OH comparison. Drawn from the Comparative Constitutions Project (CCP) codebook v4.0 adapted for U.S. subnational analysis."),
    ]
    for i, (k, v) in enumerate(doc_lines, start=1):
        c = doc.cell(row=i, column=1, value=k)
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(vertical="top")
        c = doc.cell(row=i, column=2, value=v)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        doc.row_dimensions[i].height = 40
    doc.column_dimensions["A"].width = 20
    doc.column_dimensions["B"].width = 90

    wb.save(output_path)
    print(f"Wrote codebook: {output_path}")


if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(here, "..", "input", "indicators.xlsx")
    out = os.path.normpath(out)
    build_codebook(out)
