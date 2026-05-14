"""Tests for the Excel builder.

Verifies the builder produces a structurally valid Excel:
    - All expected sheets present
    - Aggregate formulas compute to the right values
    - No formula errors
    - Handles missing/blank scores gracefully
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pytest

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

from lib.excel_builder import (
    AnalysisSection,
    Indicator,
    Score,
    State,
    build_workbook,
    load_indicators,
)


@pytest.fixture
def sample_indicators():
    """A minimal indicator set covering all three dimensions."""
    return [
        Indicator("FF01", "Fundamental Freedoms", "Rights placement", "Q1", "0-3 guidance"),
        Indicator("FF02", "Fundamental Freedoms", "Equal protection", "Q2", "0-3 guidance"),
        Indicator("CB01", "Checks and Balances", "Plural executive", "Q3", "0-3 guidance"),
        Indicator("CB02", "Checks and Balances", "AG independence", "Q4", "0-3 guidance"),
        Indicator("DP01", "Democratic Protections", "Citizen initiative", "Q5", "0-3 guidance"),
    ]


@pytest.fixture
def sample_states():
    return [
        State("SC", "South Carolina"),
        State("TN", "Tennessee"),
        State("TX", "Texas"),
        State("UT", "Utah"),
        State("MN", "Minnesota"),
        State("OH", "Ohio"),
    ]


@pytest.fixture
def sample_scores(sample_indicators, sample_states):
    """Generate a complete score set with predictable values."""
    scores = []
    for ind in sample_indicators:
        for i, state in enumerate(sample_states):
            # Use predictable scores: rotate 0-3 per state
            score = i % 4
            scores.append(
                Score(
                    indicator_id=ind.id,
                    state_code=state.code,
                    score=score,
                    constitutional_reference=f"Article I §{i+1}",
                    note=f"{state.name} note for {ind.id}.",
                )
            )
    return scores


@pytest.fixture
def sample_analysis():
    return [
        AnalysisSection("executive_summary", "This is the executive summary."),
        AnalysisSection("key_pattern", "This is the key pattern."),
        AnalysisSection("profile_SC", "SC profile."),
        AnalysisSection("profile_TN", "TN profile."),
        AnalysisSection("profile_TX", "TX profile."),
        AnalysisSection("profile_UT", "UT profile."),
        AnalysisSection("profile_MN", "MN profile."),
        AnalysisSection("profile_OH", "OH profile."),
        AnalysisSection("selection_guidance", "Selection guidance."),
    ]


def test_build_workbook_creates_file(tmp_path, sample_indicators, sample_states, sample_scores, sample_analysis):
    """The builder should produce a non-empty Excel file."""
    output = tmp_path / "out.xlsx"
    build_workbook(sample_indicators, sample_states, sample_scores, sample_analysis, output)
    assert output.exists()
    assert output.stat().st_size > 1000


def test_workbook_has_expected_sheets(tmp_path, sample_indicators, sample_states, sample_scores, sample_analysis):
    """The output should contain all required sheets in the right order."""
    output = tmp_path / "out.xlsx"
    build_workbook(sample_indicators, sample_states, sample_scores, sample_analysis, output)

    from openpyxl import load_workbook
    wb = load_workbook(output)
    expected = [
        "Read Me First",
        "Executive Summary",
        "FF — Comparative",
        "CB — Comparative",
        "DP — Comparative",
        "Tidy Data",
        "Analysis",
        "Sources",
    ]
    assert wb.sheetnames == expected


def test_tidy_data_row_count(tmp_path, sample_indicators, sample_states, sample_scores, sample_analysis):
    """Tidy data should have exactly (indicators × states) data rows."""
    output = tmp_path / "out.xlsx"
    build_workbook(sample_indicators, sample_states, sample_scores, sample_analysis, output)

    from openpyxl import load_workbook
    wb = load_workbook(output)
    ws = wb["Tidy Data"]
    # Headers in row 4, data starts at row 5
    data_rows = [r for r in ws.iter_rows(min_row=5, values_only=True) if any(c is not None for c in r)]
    expected = len(sample_indicators) * len(sample_states)
    assert len(data_rows) == expected, f"Expected {expected} rows, got {len(data_rows)}"


def test_aggregate_formulas_compute_correctly(tmp_path, sample_indicators, sample_states, sample_scores, sample_analysis):
    """The dimension aggregate formulas should sum the right scores when recalculated."""
    output = tmp_path / "out.xlsx"
    build_workbook(sample_indicators, sample_states, sample_scores, sample_analysis, output)

    # Use the project recalc script to evaluate formulas
    import subprocess
    import json

    recalc_script = HERE.parent.parent.parent.parent
    recalc = Path("/mnt/skills/public/xlsx/scripts/recalc.py")
    if not recalc.exists():
        pytest.skip("recalc.py not available in this environment")

    result = subprocess.run(
        ["python", str(recalc), str(output), "60"],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, f"recalc failed: {result.stderr}"

    info = json.loads(result.stdout)
    assert info.get("status") == "success", f"recalc found errors: {info}"
    assert info.get("total_errors", -1) == 0

    # Now verify the computed values
    from openpyxl import load_workbook
    wb = load_workbook(output, data_only=True)
    ws = wb["Executive Summary"]

    # Compute expected: for each indicator each state has score = (state_index % 4)
    # 2 FF indicators, 2 CB, 1 DP, 6 states with indices 0..5 → scores 0,1,2,3,0,1
    # Per state, the total per dimension is (n_indicators_in_dim × score_for_that_state)
    # Wait — scores are per (indicator, state), and the SAME score for ALL indicators for a given state (since score = state_index % 4).
    # So FF aggregate for state with index i = 2 * (i % 4)
    # CB aggregate = 2 * (i % 4); DP = 1 * (i % 4)

    expected_ff = [2 * (i % 4) for i in range(6)]  # [0, 2, 4, 6, 0, 2]
    expected_cb = [2 * (i % 4) for i in range(6)]
    expected_dp = [1 * (i % 4) for i in range(6)]

    # Read row 7 (FF), row 8 (CB), row 9 (DP) — cols C through H (6 states)
    for j in range(6):
        col = 3 + j  # C=3, D=4, ..., H=8
        ff_val = ws.cell(row=7, column=col).value
        cb_val = ws.cell(row=8, column=col).value
        dp_val = ws.cell(row=9, column=col).value
        assert ff_val == expected_ff[j], f"FF state {j}: expected {expected_ff[j]}, got {ff_val}"
        assert cb_val == expected_cb[j], f"CB state {j}: expected {expected_cb[j]}, got {cb_val}"
        assert dp_val == expected_dp[j], f"DP state {j}: expected {expected_dp[j]}, got {dp_val}"


def test_handles_blank_scores(tmp_path, sample_indicators, sample_states, sample_analysis):
    """The builder should not crash when scores are empty (e.g., before Claude Code runs)."""
    output = tmp_path / "out.xlsx"
    build_workbook(sample_indicators, sample_states, [], sample_analysis, output)
    assert output.exists()


def test_load_indicators_from_real_codebook():
    """The actual codebook in input/ should load and have the expected structure."""
    codebook = HERE.parent.parent / "input" / "indicators.xlsx"
    if not codebook.exists():
        pytest.skip("indicators.xlsx not yet generated; run code/00_build_codebook.py")

    indicators = load_indicators(codebook)
    assert len(indicators) >= 20, f"Expected at least 20 indicators, got {len(indicators)}"

    # Check we have all 3 dimensions
    dims = {ind.dimension for ind in indicators}
    expected_dims = {"Fundamental Freedoms", "Checks and Balances", "Democratic Protections"}
    assert dims == expected_dims, f"Dimensions: {dims}"

    # All IDs should be unique
    ids = [ind.id for ind in indicators]
    assert len(ids) == len(set(ids)), "Indicator IDs must be unique"
